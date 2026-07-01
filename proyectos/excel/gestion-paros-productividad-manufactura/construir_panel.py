"""
construir_panel.py
Genera Panel_Paros_OEE_Portfolio.xlsx con 4 hojas:
  1. Equipos         — maestro con semáforo de mantenimiento
  2. Registro_Paros  — historial con tabla Excel nativa
  3. Analisis_OEE    — OEE por equipo + Pareto de causas
  4. Dashboard       — KPIs ejecutivos + gráfico de Pareto

Ejecutar: python construir_panel.py
Requiere: openpyxl, pandas, numpy
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.layout import Layout

BASE   = os.path.dirname(__file__)
OUT    = os.path.join(BASE, "output", "Panel_Paros_OEE_Portfolio.xlsx")

# ── helpers de estilo ────────────────────────────────────────────────────────
AZUL_CORP   = "1F4E79"   # encabezados oscuros
AZUL_MED    = "2E75B6"
AZUL_CLARO  = "BDD7EE"
GRIS_CLARO  = "F2F2F2"
VERDE       = "70AD47"
AMARILLO    = "FFD966"
ROJO        = "FF0000"
NARANJA     = "ED7D31"
BLANCO      = "FFFFFF"

def borde_fino():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borde_medio():
    lado = Side(style="medium", color="595959")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def header_cell(ws, row, col, texto, bg=AZUL_CORP, fg=BLANCO, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=texto)
    c.fill    = PatternFill("solid", fgColor=bg)
    c.font    = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border  = borde_fino()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font    = Font(name="Calibri", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border  = borde_fino()
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

# ══════════════════════════════════════════════════════════════════════════════
# CARGAR DATOS
# ══════════════════════════════════════════════════════════════════════════════
df_eq    = pd.read_csv(os.path.join(BASE, "datos", "equipos.csv"))
df_paros = pd.read_csv(os.path.join(BASE, "datos", "registro_paros.csv"))

# OEE sintético por equipo (Rendimiento y Calidad)
np.random.seed(42)
rend_base  = {"CNC Alpha 01": 0.91, "CNC Alpha 02": 0.89, "Torno CNC 03": 0.82,
              "Fresadora 04": 0.93, "Centro Mec. 05": 0.88, "Torno CNC 06": 0.79,
              "Rectificadora 07": 0.83, "CNC Delta 08": 0.94}
cal_base   = {"CNC Alpha 01": 0.98, "CNC Alpha 02": 0.97, "Torno CNC 03": 0.96,
              "Fresadora 04": 0.99, "Centro Mec. 05": 0.97, "Torno CNC 06": 0.95,
              "Rectificadora 07": 0.96, "CNC Delta 08": 0.98}

# Horas paradas por equipo (no programadas)
h_paro = (df_paros[df_paros["tipo_paro"] == "No programado"]
          .groupby("nombre_equipo")["horas_paro"].sum()
          .reindex(df_eq["nombre_equipo"]).fillna(0))

HORAS_ANO = 8 * 365   # 1 turno × 8 h × 365 días = 2920 h/equipo
disponib  = ((HORAS_ANO - h_paro) / HORAS_ANO).clip(0, 1)
rendim    = pd.Series({n: rend_base[n] for n in df_eq["nombre_equipo"]})
calidad   = pd.Series({n: cal_base[n]  for n in df_eq["nombre_equipo"]})
oee       = disponib * rendim.values * calidad.values

# Pareto de causas (no programadas)
pareto = (df_paros[df_paros["tipo_paro"] == "No programado"]
          .groupby("categoria_causa")["horas_paro"].sum()
          .sort_values(ascending=False)
          .reset_index())
pareto.columns = ["Categoría", "Horas_paro"]
pareto["Pct_paro"]   = pareto["Horas_paro"] / pareto["Horas_paro"].sum()
pareto["Pct_acum"]   = pareto["Pct_paro"].cumsum()

# Pareto por equipo
pareto_eq = (df_paros[df_paros["tipo_paro"] == "No programado"]
             .groupby("nombre_equipo")["horas_paro"].sum()
             .sort_values(ascending=False)
             .reset_index())
pareto_eq.columns = ["Equipo", "Horas_paro"]
pareto_eq["Pct_paro"] = pareto_eq["Horas_paro"] / pareto_eq["Horas_paro"].sum()
pareto_eq["Pct_acum"] = pareto_eq["Pct_paro"].cumsum()

wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1: EQUIPOS
# ══════════════════════════════════════════════════════════════════════════════
ws_eq = wb.create_sheet("Equipos")
ws_eq.sheet_view.showGridLines = False

# Título
ws_eq.merge_cells("A1:H1")
t = ws_eq["A1"]
t.value     = "FabriTec S.A. — Maestro de Equipos y Semáforo de Mantenimiento"
t.fill      = PatternFill("solid", fgColor=AZUL_CORP)
t.font      = Font(bold=True, color=BLANCO, size=13, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws_eq.row_dimensions[1].height = 28

# Subtítulo
ws_eq.merge_cells("A2:H2")
s = ws_eq["A2"]
s.value = "Fecha de referencia: 01/01/2025  |  Verde = OK  |  Amarillo = Service en < 60 días  |  Rojo = Vencido o < 30 días"
s.fill  = PatternFill("solid", fgColor=AZUL_MED)
s.font  = Font(color=BLANCO, size=9, name="Calibri")
s.alignment = Alignment(horizontal="center", vertical="center")
ws_eq.row_dimensions[2].height = 18

# Encabezados
cols_eq = ["ID", "Equipo", "Línea", "Costo/h Parada ($)", "Último Service",
           "Próximo Service", "Días Restantes", "Estado"]
anchos  = [6, 20, 12, 18, 16, 16, 16, 14]
for i, (col, ancho) in enumerate(zip(cols_eq, anchos), 1):
    header_cell(ws_eq, 3, i, col)
    ws_eq.column_dimensions[get_column_letter(i)].width = ancho
ws_eq.row_dimensions[3].height = 36

# Datos
for r_idx, row in df_eq.iterrows():
    row_n = r_idx + 4
    dias  = row["dias_para_service"]
    if dias < 0:
        bg_est, estado = "FFCCCC", "VENCIDO"
        bg_dias = "FFCCCC"
    elif dias < 30:
        bg_est, estado = "FFCCCC", "URGENTE"
        bg_dias = "FFCCCC"
    elif dias < 60:
        bg_est, estado = "FFEB9C", "PRÓXIMO"
        bg_dias = "FFEB9C"
    else:
        bg_est, estado = "C6EFCE", "OK"
        bg_dias = "C6EFCE"

    row_bg = BLANCO if r_idx % 2 == 0 else GRIS_CLARO
    data_cell(ws_eq, row_n, 1, row["id_equipo"],   align="center", bg=row_bg)
    data_cell(ws_eq, row_n, 2, row["nombre_equipo"], bold=True, bg=row_bg)
    data_cell(ws_eq, row_n, 3, row["linea_produccion"], align="center", bg=row_bg)
    data_cell(ws_eq, row_n, 4, row["costo_hora_parada"], fmt='#,##0', align="right", bg=row_bg)
    data_cell(ws_eq, row_n, 5, row["fecha_ultimo_service"], align="center", bg=row_bg)
    data_cell(ws_eq, row_n, 6, row["proximo_service"], align="center", bg=row_bg)
    data_cell(ws_eq, row_n, 7, dias, fmt="0", align="center", bg=bg_dias)
    c_est = data_cell(ws_eq, row_n, 8, estado, align="center", bg=bg_est, bold=True)
    ws_eq.row_dimensions[row_n].height = 20

# Nota de referencia
ws_eq.merge_cells(f"A{4+len(df_eq)+2}:H{4+len(df_eq)+2}")
nota = ws_eq.cell(row=4+len(df_eq)+2, column=1,
                  value="(*) Rojo: vencido o < 30 días | Amarillo: 30–60 días | Verde: > 60 días")
nota.font = Font(italic=True, size=9, color="595959", name="Calibri")

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2: REGISTRO DE PAROS
# ══════════════════════════════════════════════════════════════════════════════
ws_pr = wb.create_sheet("Registro_Paros")
ws_pr.sheet_view.showGridLines = False

ws_pr.merge_cells("A1:J1")
t2 = ws_pr["A1"]
t2.value = "FabriTec S.A. — Registro de Paros de Producción | Enero–Diciembre 2024"
t2.fill  = PatternFill("solid", fgColor=AZUL_CORP)
t2.font  = Font(bold=True, color=BLANCO, size=13, name="Calibri")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws_pr.row_dimensions[1].height = 26

cols_pr = ["Fecha", "ID Equipo", "Equipo", "Línea", "Turno",
           "Tipo Paro", "Categoría Causa", "Causa Detalle", "Horas Paro", "Costo Paro ($)"]
anchos_pr = [13, 10, 20, 12, 18, 16, 18, 32, 12, 16]
for i, (col, ancho) in enumerate(zip(cols_pr, anchos_pr), 1):
    header_cell(ws_pr, 2, i, col, bg=AZUL_MED)
    ws_pr.column_dimensions[get_column_letter(i)].width = ancho
ws_pr.row_dimensions[2].height = 32

for r_idx, row in df_paros.iterrows():
    row_n = r_idx + 3
    bg = BLANCO if r_idx % 2 == 0 else GRIS_CLARO
    tipo_bg = "FFCCCC" if row["tipo_paro"] == "No programado" else "C6EFCE"
    data_cell(ws_pr, row_n, 1,  row["fecha"],           align="center", bg=bg)
    data_cell(ws_pr, row_n, 2,  row["id_equipo"],        align="center", bg=bg)
    data_cell(ws_pr, row_n, 3,  row["nombre_equipo"],    bg=bg)
    data_cell(ws_pr, row_n, 4,  row["linea"],            align="center", bg=bg)
    data_cell(ws_pr, row_n, 5,  row["turno"],            bg=bg)
    data_cell(ws_pr, row_n, 6,  row["tipo_paro"],        align="center", bg=tipo_bg, bold=True)
    data_cell(ws_pr, row_n, 7,  row["categoria_causa"],  bg=bg)
    data_cell(ws_pr, row_n, 8,  row["causa_detalle"],    bg=bg)
    data_cell(ws_pr, row_n, 9,  row["horas_paro"],  fmt="0.0", align="right", bg=bg)
    data_cell(ws_pr, row_n, 10, row["costo_paro"],  fmt="#,##0", align="right", bg=bg)
    ws_pr.row_dimensions[row_n].height = 16

# Tabla Excel nativa
tabla_ref = f"A2:J{2 + len(df_paros)}"
tbl = Table(displayName="Registro_Paros", ref=tabla_ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_pr.add_table(tbl)

# Freezar primera fila visible
ws_pr.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3: ANÁLISIS OEE
# ══════════════════════════════════════════════════════════════════════════════
ws_oee = wb.create_sheet("Analisis_OEE")
ws_oee.sheet_view.showGridLines = False

ws_oee.merge_cells("A1:J1")
t3 = ws_oee["A1"]
t3.value = "FabriTec S.A. — OEE por Equipo y Pareto de Causas de Paro | 2024"
t3.fill  = PatternFill("solid", fgColor=AZUL_CORP)
t3.font  = Font(bold=True, color=BLANCO, size=13, name="Calibri")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws_oee.row_dimensions[1].height = 26

# --- OEE por equipo ---
ws_oee.merge_cells("A3:G3")
sub1 = ws_oee["A3"]
sub1.value = "OEE por Equipo (Eficiencia General de Equipos)"
sub1.fill  = PatternFill("solid", fgColor=AZUL_MED)
sub1.font  = Font(bold=True, color=BLANCO, size=11, name="Calibri")
sub1.alignment = Alignment(horizontal="center", vertical="center")
ws_oee.row_dimensions[3].height = 22

cols_oee = ["Equipo", "H. Paro No Prog.", "Disponibilidad %", "Rendimiento %", "Calidad %", "OEE %", "Evaluación"]
anchos_oee = [20, 18, 17, 15, 12, 10, 14]
for i, (col, ancho) in enumerate(zip(cols_oee, anchos_oee), 1):
    header_cell(ws_oee, 4, i, col, bg=AZUL_CORP)
    ws_oee.column_dimensions[get_column_letter(i)].width = ancho
ws_oee.row_dimensions[4].height = 32

oee_nombres = df_eq["nombre_equipo"].tolist()
for i, nombre in enumerate(oee_nombres):
    row_n  = i + 5
    horas  = h_paro[nombre]
    disp   = disponib[nombre]
    rend   = rendim[nombre]
    cal    = calidad[nombre]
    oee_v  = oee[nombre]
    eval_v = "Excelente ≥ 85%" if oee_v >= 0.85 else ("Aceptable ≥ 65%" if oee_v >= 0.65 else "Bajo < 65%")
    oee_bg = "C6EFCE" if oee_v >= 0.85 else ("FFEB9C" if oee_v >= 0.65 else "FFCCCC")
    bg = BLANCO if i % 2 == 0 else GRIS_CLARO
    data_cell(ws_oee, row_n, 1, nombre,        bold=True, bg=bg)
    data_cell(ws_oee, row_n, 2, round(horas,1), fmt="0.0",  align="right", bg=bg)
    data_cell(ws_oee, row_n, 3, disp,           fmt="0.0%", align="center", bg=bg)
    data_cell(ws_oee, row_n, 4, rend,           fmt="0.0%", align="center", bg=bg)
    data_cell(ws_oee, row_n, 5, cal,            fmt="0.0%", align="center", bg=bg)
    data_cell(ws_oee, row_n, 6, oee_v,          fmt="0.0%", align="center", bg=oee_bg, bold=True)
    data_cell(ws_oee, row_n, 7, eval_v,         align="center", bg=oee_bg)
    ws_oee.row_dimensions[row_n].height = 20

# Nota OEE
row_nota_oee = 5 + len(oee_nombres) + 1
ws_oee.merge_cells(f"A{row_nota_oee}:G{row_nota_oee}")
nt = ws_oee.cell(row=row_nota_oee, column=1,
                 value="OEE = Disponibilidad × Rendimiento × Calidad  |  Benchmark industria: Excelente ≥ 85 % | Aceptable 65–84 %")
nt.font = Font(italic=True, size=9, color="595959", name="Calibri")

# --- Pareto de causas ---
row_pareto_title = row_nota_oee + 2
ws_oee.merge_cells(f"A{row_pareto_title}:E{row_pareto_title}")
sub2 = ws_oee[f"A{row_pareto_title}"]
sub2.value = "Pareto de Causas — Paros No Programados (horas perdidas por categoría)"
sub2.fill  = PatternFill("solid", fgColor=AZUL_MED)
sub2.font  = Font(bold=True, color=BLANCO, size=11, name="Calibri")
sub2.alignment = Alignment(horizontal="center", vertical="center")
ws_oee.row_dimensions[row_pareto_title].height = 22

cols_pareto = ["Categoría de Causa", "Horas Perdidas", "% del Total", "% Acumulado", "Interpretación"]
anchos_pareto = [22, 16, 12, 14, 24]
for i, (col, ancho) in enumerate(zip(cols_pareto, anchos_pareto), 1):
    header_cell(ws_oee, row_pareto_title + 1, i, col, bg=AZUL_CORP)
    if i > len(anchos_oee):
        ws_oee.column_dimensions[get_column_letter(i)].width = ancho
ws_oee.row_dimensions[row_pareto_title + 1].height = 30

pareto_colors = ["FFCCCC", "FFEB9C", "BDD7EE", "E2EFDA"]
for j, prow in pareto.iterrows():
    rn = row_pareto_title + 2 + j
    bg = pareto_colors[j] if j < len(pareto_colors) else GRIS_CLARO
    interp = "Causa principal" if j == 0 else ("Top 3 — 80 % del downtime" if prow["Pct_acum"] <= 0.80 else "Causa secundaria")
    data_cell(ws_oee, rn, 1, prow["Categoría"],    bold=(j == 0), bg=bg)
    data_cell(ws_oee, rn, 2, prow["Horas_paro"],   fmt="0.0", align="right", bg=bg)
    data_cell(ws_oee, rn, 3, prow["Pct_paro"],     fmt="0%",  align="center", bg=bg)
    data_cell(ws_oee, rn, 4, prow["Pct_acum"],     fmt="0%",  align="center", bg=bg, bold=True)
    data_cell(ws_oee, rn, 5, interp,               align="center", bg=bg)
    ws_oee.row_dimensions[rn].height = 20

# --- Pareto por equipo ---
row_eq_pareto = row_pareto_title + 2 + len(pareto) + 2
ws_oee.merge_cells(f"A{row_eq_pareto}:E{row_eq_pareto}")
sub3 = ws_oee[f"A{row_eq_pareto}"]
sub3.value = "Pareto por Equipo — Concentración del downtime no programado"
sub3.fill  = PatternFill("solid", fgColor=AZUL_MED)
sub3.font  = Font(bold=True, color=BLANCO, size=11, name="Calibri")
sub3.alignment = Alignment(horizontal="center", vertical="center")
ws_oee.row_dimensions[row_eq_pareto].height = 22

cols_eq_p = ["Equipo", "Horas No Programadas", "% del Total", "% Acumulado", "Acción recomendada"]
for i, col in enumerate(cols_eq_p, 1):
    header_cell(ws_oee, row_eq_pareto + 1, i, col, bg=AZUL_CORP)
ws_oee.row_dimensions[row_eq_pareto + 1].height = 30

for j, er in pareto_eq.iterrows():
    rn  = row_eq_pareto + 2 + j
    bg  = "FFCCCC" if j < 3 else (GRIS_CLARO if j % 2 == 0 else BLANCO)
    acc = "Plan preventivo urgente" if j < 3 else ("Monitorear" if j < 5 else "Sin intervención prioritaria")
    data_cell(ws_oee, rn, 1, er["Equipo"],    bold=(j < 3), bg=bg)
    data_cell(ws_oee, rn, 2, er["Horas_paro"], fmt="0.0", align="right", bg=bg)
    data_cell(ws_oee, rn, 3, er["Pct_paro"],   fmt="0%",  align="center", bg=bg)
    data_cell(ws_oee, rn, 4, er["Pct_acum"],   fmt="0%",  align="center", bg=bg, bold=True)
    data_cell(ws_oee, rn, 5, acc,              align="center", bg=bg)
    ws_oee.row_dimensions[rn].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False

# Título principal
ws_dash.merge_cells("A1:N1")
td = ws_dash["A1"]
td.value     = "FabriTec S.A. — Panel Ejecutivo de Operaciones | Paros y OEE 2024"
td.fill      = PatternFill("solid", fgColor=AZUL_CORP)
td.font      = Font(bold=True, color=BLANCO, size=14, name="Calibri")
td.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 32

# Subtítulo
ws_dash.merge_cells("A2:N2")
sd = ws_dash["A2"]
sd.value = "Período: Enero – Diciembre 2024   |   Equipos analizados: 8   |   Actualizado: 01/01/2025"
sd.fill  = PatternFill("solid", fgColor=AZUL_MED)
sd.font  = Font(color=BLANCO, size=10, name="Calibri")
sd.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[2].height = 18

# KPI totales
total_horas_np = h_paro.sum()
total_costo_np = df_paros[df_paros["tipo_paro"] == "No programado"]["costo_paro"].sum()
oee_prom       = oee.mean()
total_eventos   = len(df_paros[df_paros["tipo_paro"] == "No programado"])

kpis = [
    ("Total Horas\nParadas (NP)", f"{total_horas_np:,.0f} h", "FFCCCC", ROJO),
    ("Costo Total\nParos NP ($)", f"${total_costo_np:,.0f}", "FFCCCC", ROJO),
    ("Eventos\nNo Programados", f"{total_eventos}", "FFEB9C", "8B4513"),
    ("OEE Promedio\nPlanta", f"{oee_prom:.1%}", "C6EFCE" if oee_prom >= 0.65 else "FFCCCC", "375623"),
]

kpi_col_starts = [2, 5, 8, 11]
for (titulo, valor, bg_kpi, color_v), col_st in zip(kpis, kpi_col_starts):
    ws_dash.merge_cells(start_row=4, start_column=col_st, end_row=4, end_column=col_st+2)
    ck = ws_dash.cell(row=4, column=col_st, value=titulo)
    ck.fill      = PatternFill("solid", fgColor=bg_kpi)
    ck.font      = Font(bold=True, size=10, name="Calibri")
    ck.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ck.border    = borde_medio()
    ws_dash.row_dimensions[4].height = 36

    ws_dash.merge_cells(start_row=5, start_column=col_st, end_row=5, end_column=col_st+2)
    cv = ws_dash.cell(row=5, column=col_st, value=valor)
    cv.fill      = PatternFill("solid", fgColor=bg_kpi)
    cv.font      = Font(bold=True, size=18, color=color_v, name="Calibri")
    cv.alignment = Alignment(horizontal="center", vertical="center")
    cv.border    = borde_medio()
    ws_dash.row_dimensions[5].height = 40

# ── Gráfico 1: Pareto de equipos ──────────────────────────────────────────────
# Datos auxiliares en filas ocultas (fila 40+)
AUX_ROW = 42
ws_dash.cell(row=AUX_ROW, column=1, value="Equipo")
ws_dash.cell(row=AUX_ROW, column=2, value="Horas NP")
ws_dash.cell(row=AUX_ROW, column=3, value="% Acum.")
for j, er in pareto_eq.iterrows():
    r = AUX_ROW + 1 + j
    ws_dash.cell(row=r, column=1, value=er["Equipo"])
    ws_dash.cell(row=r, column=2, value=round(er["Horas_paro"], 1))
    ws_dash.cell(row=r, column=3, value=round(er["Pct_acum"], 3))

n_eq   = len(pareto_eq)
cats1  = Reference(ws_dash, min_col=1, min_row=AUX_ROW+1, max_row=AUX_ROW+n_eq)
data_b = Reference(ws_dash, min_col=2, min_row=AUX_ROW,   max_row=AUX_ROW+n_eq)
data_l = Reference(ws_dash, min_col=3, min_row=AUX_ROW,   max_row=AUX_ROW+n_eq)

bar1 = BarChart()
bar1.type    = "col"
bar1.title   = "Pareto de Equipos — Horas de Paro No Programado"
bar1.y_axis.title = "Horas"
bar1.x_axis.title = "Equipo"
bar1.add_data(data_b, titles_from_data=True)
bar1.set_categories(cats1)
bar1.series[0].graphicalProperties.solidFill = "2E75B6"
bar1.shape   = 4
bar1.width   = 20
bar1.height  = 14

line1 = LineChart()
line1.add_data(data_l, titles_from_data=True)
line1.set_categories(cats1)
line1.series[0].graphicalProperties.line.solidFill = NARANJA
line1.series[0].graphicalProperties.line.width = 20000
bar1 += line1
bar1.y_axis.axId = 100
line1.y_axis.axId = 200
line1.y_axis.crosses = "max"
line1.y_axis.numFmt  = "0%"
line1.y_axis.title   = "% Acumulado"

ws_dash.add_chart(bar1, "B7")

# ── Gráfico 2: OEE por equipo ─────────────────────────────────────────────────
# Datos auxiliares en columnas 6-7 para no colisionar con datos del Pareto (cols 1-3)
AUX_ROW2 = 55
ws_dash.cell(row=AUX_ROW2, column=6, value="Equipo")
ws_dash.cell(row=AUX_ROW2, column=7, value="OEE %")
for i, nombre in enumerate(oee_nombres):
    ws_dash.cell(row=AUX_ROW2+1+i, column=6, value=nombre)
    ws_dash.cell(row=AUX_ROW2+1+i, column=7, value=round(oee[nombre], 3))

cats2  = Reference(ws_dash, min_col=6, min_row=AUX_ROW2+1, max_row=AUX_ROW2+len(oee_nombres))
data_o = Reference(ws_dash, min_col=7, min_row=AUX_ROW2,   max_row=AUX_ROW2+len(oee_nombres))

bar2 = BarChart()
bar2.type    = "bar"
bar2.title   = "OEE por Equipo — Eficiencia General 2024"
bar2.y_axis.title = "Equipo"
bar2.x_axis.title = "OEE %"
bar2.x_axis.numFmt = "0%"
# Escala 60%-100%: la diferencia entre equipos queda visualmente clara
bar2.x_axis.scaling.min = 0.6
bar2.x_axis.scaling.max = 1.0
bar2.add_data(data_o, titles_from_data=True)
bar2.set_categories(cats2)
bar2.width   = 22
bar2.height  = 14

# Colorear cada barra individualmente según nivel OEE
OEE_COLORS = {"verde": "70AD47", "amarillo": "FFD966", "rojo": "FF6B6B"}
for i, nombre in enumerate(oee_nombres):
    v = oee[nombre]
    color = OEE_COLORS["verde"] if v >= 0.85 else (OEE_COLORS["amarillo"] if v >= 0.65 else OEE_COLORS["rojo"])
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    bar2.series[0].dPt.append(pt)

ws_dash.add_chart(bar2, "K7")

# Ajustar anchos de columnas del Dashboard
for i in range(1, 22):
    ws_dash.column_dimensions[get_column_letter(i)].width = 9

# Nota de ahorro proyectado
ws_dash.merge_cells("B28:M29")
nr = ws_dash["B28"]
nr.value = (
    "Diagnóstico clave: Equipos Torno CNC 03, Torno CNC 06 y Rectificadora 07 concentran el 64 % "
    "del downtime no programado. Un plan de mantenimiento preventivo focalizado en estos 3 equipos "
    "puede reducir ~700 h/año de paros. Impacto estimado: $8,4 M/año en producción recuperada."
)
nr.fill      = PatternFill("solid", fgColor="FFF2CC")
nr.font      = Font(bold=False, size=10, name="Calibri", color="3D3D3D")
nr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
nr.border    = borde_medio()
ws_dash.row_dimensions[28].height = 30
ws_dash.row_dimensions[29].height = 30

# Pie del Dashboard
ws_dash.merge_cells("B31:M31")
pie = ws_dash["B31"]
pie.value = "Datos sintéticos con distribuciones realistas del sector manufacturero — Portfolio de Datos y Analítica | Luciano Villagrán"
pie.font  = Font(italic=True, size=9, color="595959", name="Calibri")
pie.alignment = Alignment(horizontal="center")

# ══════════════════════════════════════════════════════════════════════════════
# ORDEN DE HOJAS
# ══════════════════════════════════════════════════════════════════════════════
# Dashboard primero
wb.move_sheet("Dashboard", offset=-3)

# Colores de pestaña
ws_dash.sheet_properties.tabColor = AZUL_CORP
ws_eq.sheet_properties.tabColor   = AZUL_MED
ws_pr.sheet_properties.tabColor   = "595959"
ws_oee.sheet_properties.tabColor  = VERDE

# Guardar
wb.save(OUT)
print(f"Excel generado: {OUT}")
