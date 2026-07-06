# -*- coding: utf-8 -*-
"""
generar_dashboard.py
Genera output/dashboard_inventario_fmcg.png: una vista previa del panel de control,
leyendo directamente la hoja 'Análisis' de Inventario_FMCG_Portfolio.xlsx.
Pensado para que se entienda SIN abrir la planilla (útil en el README).

Reglas de diseño (skill dashboards-python): título = conclusión, colores semáforo,
separadores de miles, una idea por gráfico.

Ejecutar:  python generar_dashboard.py
Requiere:  pandas, matplotlib, openpyxl
"""
import os
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec

plt.rcParams["mathtext.default"] = "regular"

BASE = os.path.dirname(__file__)
XLSX = os.path.join(BASE, "Inventario_FMCG_Portfolio.xlsx")
OUT  = os.path.join(BASE, "output", "dashboard_inventario_fmcg.png")

AZUL, AZULM, VERDE = "#1F4E79", "#2E75B6", "#70AD47"
AMAR, ROJO, NARANJA, MORADO = "#FFC000", "#C00000", "#ED7D31", "#7030A0"

def miles(x):
    return f"{x:,.0f}".replace(",", ".")

# ── Leer la hoja 'Análisis' (encabezado en la fila 4) ────────────────────────
import openpyxl
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Análisis"]
headers = [str(ws.cell(4, c).value).replace("\n", " ") if ws.cell(4, c).value else f"col{c}"
           for c in range(1, 17)]
rows = [[ws.cell(r, c).value for c in range(1, 17)]
        for r in range(5, ws.max_row + 1) if ws.cell(r, 1).value is not None]
df = pd.DataFrame(rows, columns=headers)

cats_validas = ["Alimentos Secos", "Limpieza del Hogar", "Higiene Personal"]
df = df[df["Categoría"].isin(cats_validas)].copy()
df["Capital Inmovilizado"] = pd.to_numeric(df["Capital Inmovilizado"], errors="coerce")

def estado(s):
    if pd.isna(s): return None
    for k in ["QUIEBRE", "CRÍTICO", "NORMAL", "SOBRESTOCK", "STOCK MUERTO"]:
        if k in str(s): return k
    return None
df["estado"] = df["Estado Cobertura"].map(estado)

counts = df["estado"].value_counts()
n_quiebre = int(counts.get("QUIEBRE", 0))
n_muerto  = int(counts.get("STOCK MUERTO", 0))
cap_muerto = df.loc[df["estado"] == "STOCK MUERTO", "Capital Inmovilizado"].sum()
n_sku = len(df)
abc = df["Clasif. ABC"].value_counts().reindex(["A", "B", "C"]).fillna(0)

# ── Figura ───────────────────────────────────────────────────────────────────
fig = plt.figure(figsize=(13, 8.2)); fig.patch.set_facecolor("white")
gs = GridSpec(3, 4, figure=fig, height_ratios=[0.62, 1.15, 1.15], hspace=0.6, wspace=0.42,
              left=0.06, right=0.965, top=0.9, bottom=0.09)

fig.text(0.06, 0.955, "Distribuidora FMCG — Panel de Control de Inventario (75 productos)",
         fontsize=17, fontweight="bold", color=AZUL, ha="left")
fig.text(0.06, 0.925, "Se detectaron 7 productos por quebrar stock y USD 22.473 de capital atrapado en mercadería que no rota",
         fontsize=11, color="#7F7F7F", ha="left")

kpis = [("Productos\nanalizados", f"{n_sku}", "#EAF1F8", AZUL),
        ("Por quebrar stock\n(venta en riesgo)", f"{n_quiebre}", "#FDECEA", ROJO),
        ("Sin rotación\n(stock muerto)", f"{n_muerto}", "#F1EAF7", MORADO),
        ("Capital atrapado\nidentificado", f"USD {miles(cap_muerto)}", "#FEF6E7", "#8B5E00")]
for i,(t,v,bg,cv) in enumerate(kpis):
    ax = fig.add_subplot(gs[0, i]); ax.axis("off")
    ax.add_patch(plt.Rectangle((0,0),1,1, facecolor=bg, edgecolor="#D9D9D9", lw=1, transform=ax.transAxes))
    ax.text(0.5,0.72,t, ha="center", va="center", fontsize=9.3, color="#404040", transform=ax.transAxes)
    ax.text(0.5,0.33,v, ha="center", va="center", fontsize=18, fontweight="bold", color=cv, transform=ax.transAxes)

estado_order = ["QUIEBRE","CRÍTICO","NORMAL","SOBRESTOCK","STOCK MUERTO"]
estado_lbl = {"QUIEBRE":"Quiebre\n(sin stock)","CRÍTICO":"Crítico","NORMAL":"Normal",
              "SOBRESTOCK":"Sobre-\nstock","STOCK MUERTO":"Stock muerto\n(no rota)"}
estado_col = {"QUIEBRE":ROJO,"CRÍTICO":NARANJA,"NORMAL":VERDE,"SOBRESTOCK":AMAR,"STOCK MUERTO":MORADO}
ax1 = fig.add_subplot(gs[1, :2]); est_vals = [int(counts.get(e,0)) for e in estado_order]
bars = ax1.bar(range(len(estado_order)), est_vals, color=[estado_col[e] for e in estado_order], edgecolor="white", zorder=2)
ax1.set_xticks(range(len(estado_order))); ax1.set_xticklabels([estado_lbl[e] for e in estado_order], fontsize=8.6)
ax1.set_ylabel("Cantidad de productos", fontsize=9.5)
ax1.set_title("Semáforo de stock — 7 productos a punto de quebrar y 12 sin rotación",
              fontsize=10.5, fontweight="bold", pad=8, loc="left")
ax1.yaxis.grid(True, ls="--", alpha=0.35, zorder=0); ax1.set_axisbelow(True)
for s in ("top","right"): ax1.spines[s].set_visible(False)
for b,v in zip(bars, est_vals): ax1.text(b.get_x()+b.get_width()/2, v+0.4, str(v), ha="center", fontsize=9, fontweight="bold")

ax2 = fig.add_subplot(gs[1, 2:]); abc_col = {"A":AZUL,"B":AZULM,"C":"#BDD7EE"}
bars2 = ax2.bar(range(3), abc.values, color=[abc_col[k] for k in ["A","B","C"]], edgecolor="white", zorder=2)
ax2.set_xticks(range(3)); ax2.set_xticklabels(["A\n(los que más\nfacturan)","B\n(intermedios)","C\n(los que menos\nfacturan)"], fontsize=8.6)
ax2.set_ylabel("Cantidad de productos", fontsize=9.5)
ax2.set_title("Clasificación ABC — dónde poner el foco de gestión", fontsize=10.5, fontweight="bold", pad=8, loc="left")
ax2.yaxis.grid(True, ls="--", alpha=0.35, zorder=0); ax2.set_axisbelow(True)
for s in ("top","right"): ax2.spines[s].set_visible(False)
for b,v in zip(bars2, abc.values): ax2.text(b.get_x()+b.get_width()/2, v+0.4, f"{int(v)}", ha="center", fontsize=9, fontweight="bold")

ax3 = fig.add_subplot(gs[2, :]); ax3.axis("off")
ax3.add_patch(plt.Rectangle((0,0),1,1, facecolor="#F0F6FF", edgecolor="#A9C7E8", lw=1.2, transform=ax3.transAxes))
diag = ("QUÉ RESOLVIÓ ESTE PANEL\n"
        "• Antes: 4–5 horas por mes armando el reporte a mano y sin forma de ver qué producto estaba por faltar.\n"
        "• Después: el estado de los 75 productos se ve de un vistazo y el reporte se arma en ~20 minutos (−93 % de tiempo).\n"
        "• Se identificaron 7 productos por quebrar (venta en riesgo) y USD 22.473 de capital atrapado en mercadería sin rotación.\n"
        "  Cifras sobre datos sintéticos del sector FMCG; capital expresado en dólares (USD).")
ax3.text(0.02,0.5, diag, ha="left", va="center", fontsize=9.6, color="#243B53", transform=ax3.transAxes, linespacing=1.5)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
fig.savefig(OUT, dpi=145, facecolor="white", bbox_inches="tight"); plt.close(fig)
print("Dashboard generado:", OUT)
